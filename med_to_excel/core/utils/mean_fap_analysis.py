from PySide6.QtGui import QClipboard
import openpyxl
import statistics
import math
from med_to_excel.core.utils.recorrence import remover_data
from med_to_excel.core.utils.error_log import set_errors, SEQUENCE_RESPONSE_ERROR



def clear_data(string):

    row_data = string.split()
    if ',' in string:
        temp = []
        for i in row_data:
            temp.append(i.replace(',', '.'))
        row_data = temp
    if '/' in string:
        row_data = remover_data(row_data)
    
    temp = []
    for i in row_data:
        if '0.000' != i:
            temp.append(i)

    row_data = temp
    return row_data

def separate_data(time_data): 
        all_latency = []
        all_f2s = []
        all_s2t = []
        for i in time_data:
            if i[-3:] == '001':
                temp = i[:-3]
                temp = temp.replace('.', '')
                all_latency.append(int(temp))
            elif i[-3:] == '002':
                temp = i[:-3]
                temp = temp.replace('.', '')
                all_f2s.append(int(temp))
            elif i[-3:] == '003':
                temp = i[:-3]
                temp = temp.replace('.', '')
                all_s2t.append(int(temp))
        
        return (all_latency, all_f2s, all_s2t)

def calc_latency(time_data, consequence_data, individual, name):
    cb = QClipboard()
    row_time_data = clear_data(time_data)
    row_consequence_data = clear_data(consequence_data)

    if individual:
        time_data_separated = []
        for i in row_time_data:
            time_data_separated.append(int(i.split('.')[0]))
    else:
        time_data_separated = separate_data(row_time_data)
    
    reinforce_counter = 1
    latency_by_reinforce = {}
    f2s_by_reinforce = {}
    s2t_by_reinforce = {}

    for trial in range(len(row_consequence_data)):
        if reinforce_counter in latency_by_reinforce.keys():
            if individual:
                latency_by_reinforce[reinforce_counter].append(time_data_separated[trial])
            else:
                try:
                    latency_by_reinforce[reinforce_counter].append(time_data_separated[0][trial])
                    f2s_by_reinforce[reinforce_counter].append(time_data_separated[1][trial])
                    s2t_by_reinforce[reinforce_counter].append(time_data_separated[2][trial])
                except IndexError:
                    pass
        else:
            if individual:
                latency_by_reinforce[reinforce_counter] = [time_data_separated[trial]]
            else:
                try:
                    latency_by_reinforce[reinforce_counter] = [time_data_separated[0][trial]]
                    f2s_by_reinforce[reinforce_counter] = [time_data_separated[1][trial]]
                    s2t_by_reinforce[reinforce_counter] = [time_data_separated[2][trial]]
                except IndexError:
                    pass

        if row_consequence_data[trial][0] == '1':
            reinforce_counter += 1
    
    wb = openpyxl.Workbook() # Create a excel Workbook 
    ws = wb.active # Create a excel worksheet
    ws.title = "Latencia"
    if individual:
        ws.append(["Latência da primeira, segunda, penúltima e última resposta"])
    else:
        ws.append(["Latência da primeira, segunda, penúltima e última sequência"])
    ws.append(["Reforço", "Primeira", "Segunda", "Penúltima", "Ultima"])
    all_relevant_latency = [],[],[],[]
    for i in latency_by_reinforce:
        temp_index_of = [0, 1, -2, -1]
        temp_cels = [f"{i}",]
        for ii in range(len(temp_index_of)):
            try:
                all_relevant_latency[ii].append(latency_by_reinforce[i][temp_index_of[ii]])
                temp_cels.append(latency_by_reinforce[i][temp_index_of[ii]])
            except IndexError:
                pass
        # all_relevant_latency[1].append(latency_by_reinforce[i][1])
        # all_relevant_latency[2].append(latency_by_reinforce[i][-2])
        # all_relevant_latency[3].append(latency_by_reinforce[i][-1])
       
        ws.append(temp_cels)

    try:
        ws.append(["Média", statistics.mean(all_relevant_latency[0]), statistics.mean(all_relevant_latency[1]), statistics.mean(all_relevant_latency[2]), statistics.mean(all_relevant_latency[3])])
    except statistics.StatisticsError:
        set_errors(SEQUENCE_RESPONSE_ERROR)
    ws.append([])
    ws.append([])

    if not individual:
        ws2 = wb.create_sheet("Primeira para Segunda") # Create a excel worksheet
        ws2.append(["Tempo entre a primeira e a segunda resposta da primeira, segunda, penúltima e última sequência"])
        ws2.append(["Reforço", "Primeira", "Segunda", "Penúltima", "Ultima"])
        all_relevant_f2s = [],[],[],[]
        for i in f2s_by_reinforce:
            temp_index_of = [0, 1, -2, -1]
            temp_cels = [f"{i}",]
            for ii in range(len(temp_index_of)):
                try:
                    all_relevant_f2s[ii].append(f2s_by_reinforce[i][temp_index_of[ii]])
                    temp_cels.append(f2s_by_reinforce[i][temp_index_of[ii]])
                except IndexError:
                    pass
    
            ws2.append(temp_cels)

        ws2.append(["Média", statistics.mean(all_relevant_f2s[0]), statistics.mean(all_relevant_f2s[1]), statistics.mean(all_relevant_f2s[2]), statistics.mean(all_relevant_f2s[3])])
        ws2.append([])
        ws2.append([])


        ws3 = wb.create_sheet("Segunda para Terceira") # Create a excel worksheet
        ws3.append(["Tempo entre a primeira e a segunda resposta da primeira, segunda, penúltima e última sequência"])
        ws3.append(["Reforço", "Primeira", "Segunda", "Penúltima", "Ultima"])
        all_relevant_s2t = [],[],[],[]
        for i in s2t_by_reinforce:
            temp_index_of = [0, 1, -2, -1]
            temp_cels = [f"{i}",]
            for ii in range(len(temp_index_of)):
                try:
                    all_relevant_s2t[ii].append(s2t_by_reinforce[i][temp_index_of[ii]])
                    temp_cels.append(s2t_by_reinforce[i][temp_index_of[ii]])
                except IndexError:
                    pass
        
            ws3.append(temp_cels)

        ws3.append(["Média", statistics.mean(all_relevant_s2t[0]), statistics.mean(all_relevant_s2t[1]), statistics.mean(all_relevant_s2t[2]), statistics.mean(all_relevant_s2t[3])])
        ws3.append([])
        ws3.append([])

    # Creating table with data of corrects sequences

    correct_latency = [[]]
    correct_f2s = [[]]
    correct_s2t = [[]]
    reinforce_counter = 0

    for i in range(len(row_consequence_data)):
        if row_consequence_data[i][0] == '1':
            if individual:
                correct_latency[reinforce_counter].append(time_data_separated[i])
            else:
                try:
                    correct_latency[reinforce_counter].append(time_data_separated[0][i])
                    correct_f2s[reinforce_counter].append(time_data_separated[1][i])
                    correct_s2t[reinforce_counter].append(time_data_separated[2][i])
                except IndexError:
                    pass
            reinforce_counter += 1
        if row_consequence_data[i][0] == '2':
            try:
                if individual:
                    correct_latency[reinforce_counter].append(time_data_separated[i])
                else:
                    correct_latency[reinforce_counter].append(time_data_separated[0][i])
                    correct_f2s[reinforce_counter].append(time_data_separated[1][i])
                    correct_s2t[reinforce_counter].append(time_data_separated[2][i])
            except IndexError:
                if individual:
                    correct_latency.append([time_data_separated[i]])
                else:
                    correct_latency.append([time_data_separated[0][i]])
                    correct_f2s.append([time_data_separated[1][i]])
                    correct_s2t.append([time_data_separated[2][i]])



    if individual:
        ws.append(["Latência das respostas corretas"])
        ws.append(["Reforço", "Primeira", "Segunda", "Terceira", "Quarta", "Quinta", "Sexta", "Setima", "Oitava", "Nona", "Decima", "Decima Primeira", "Decima Segunda"])
    else:
        ws.append(["Latência das sequências corretas"])
        ws.append(["Reforço", "Primeira", "Segunda", "Terceira", "Quarta"])

        ws2.append(["tempo entre a primeira e a segunda resposta das sequências corretas"])
        ws2.append(["Reforço", "Primeira", "Segunda", "Terceira", "Quarta"])

        ws3.append(["tempo entre a segunda e a terceira resposta das sequências corretas"])
        ws3.append(["Reforço", "Primeira", "Segunda", "Terceira", "Quarta"])

    for i in range(len(correct_latency)):
        if individual:
            try:
                ws.append([f"{i+1}", correct_latency[i][0], correct_latency[i][1], correct_latency[i][2], correct_latency[i][3], correct_latency[i][4], correct_latency[i][5], correct_latency[i][6], correct_latency[i][7], correct_latency[i][8], correct_latency[i][9], correct_latency[i][10], correct_latency[i][11]])
            except IndexError:
                temp = [f"{i+1}", ]
                for cor in correct_latency[i]:
                    temp.append(cor)
                ws.append(temp)
                
        else:
            try:
                ws2.append([f"{i+1}", correct_f2s[i][0], correct_f2s[i][1], correct_f2s[i][2], correct_f2s[i][3]])
                ws3.append([f"{i+1}", correct_s2t[i][0], correct_s2t[i][1], correct_s2t[i][2], correct_s2t[i][3]])
                ws.append([f"{i+1}", correct_latency[i][0], correct_latency[i][1], correct_latency[i][2], correct_latency[i][3]])
            except IndexError:
                temp = [f"{i+1}", ]
                for cor in correct_latency[i]:
                    temp.append(cor)
                ws.append(temp)

                temp = [f"{i+1}", ]
                for cor in correct_f2s[i]:
                    temp.append(cor)
                ws2.append(temp)

                temp = [f"{i+1}", ]
                for cor in correct_s2t[i]:
                    temp.append(cor)
                ws3.append(temp)

    if individual:
        # ws.append(["Média", statistics.mean([x[0] for x in correct_latency]), statistics.mean([x[1] for x in correct_latency]), statistics.mean([x[2] for x in correct_latency]), statistics.mean([x[3] for x in correct_latency]), statistics.mean([x[4] for x in correct_latency]), statistics.mean([x[5] for x in correct_latency]), statistics.mean([x[6] for x in correct_latency]), statistics.mean([x[7] for x in correct_latency]), statistics.mean([x[8] for x in correct_latency]), statistics.mean([x[9] for x in correct_latency]), statistics.mean([x[10] for x in correct_latency]), statistics.mean([x[11] for x in correct_latency])])
        means_latency = [[] for x in range(12)]
        for reinf in correct_latency:
            for cor in range(len(reinf)):
                means_latency[cor].append(reinf[cor])
        temp = ["Média:",]
        try:
            for cel in means_latency:
                    temp.append(statistics.mean(cel))
        except statistics.StatisticsError:
            set_errors(SEQUENCE_RESPONSE_ERROR)
        ws.append(temp)

    else:
        def append_mean(list_values):
            list_first, list_second, list_penultimate, list_last = [], [], [], []
            for i in range(len(list_values)):
                try:
                    list_first.append(list_values[i][0])
                except IndexError:
                    pass
                try:
                    list_second.append(list_values[i][1])
                except IndexError:
                    pass
                try:
                    list_penultimate.append(list_values[i][2])
                except IndexError:
                    pass
                try:
                    list_last.append(list_values[i][3])
                except IndexError:
                    pass

            return list_first, list_second, list_penultimate, list_last
        
        list_values = append_mean(correct_latency)
        ws.append(["Média", statistics.mean(list_values[0]), statistics.mean(list_values[1]), statistics.mean(list_values[2]), statistics.mean(list_values[3])])
        list_values = append_mean(correct_f2s)
        ws2.append(["Média", statistics.mean(list_values[0]), statistics.mean(list_values[1]), statistics.mean(list_values[2]), statistics.mean(list_values[3])])
        list_values = append_mean(correct_s2t)
        ws3.append(["Média", statistics.mean(list_values[0]), statistics.mean(list_values[1]), statistics.mean(list_values[2]), statistics.mean(list_values[3])])


    # Coping data to excel
    wb.save(f"./planilhas/{name}.xlsx")


def calc_sequence_duration(time_data, consequence_data, name):
    cb = QClipboard()
    row_time_data = clear_data(time_data)
    row_consequence_data = clear_data(consequence_data)

    time_data_separated = []
    for i in row_time_data:
        time_data_separated.append(int(i.split('.')[0]))
   
    reinforce_counter = 1
    duration_by_reinforce = {}

    for trial in range(len(row_consequence_data)):
        if reinforce_counter in duration_by_reinforce.keys():
            duration_by_reinforce[reinforce_counter].append(time_data_separated[trial])
        else:
            duration_by_reinforce[reinforce_counter] = [time_data_separated[trial]]
    
        if row_consequence_data[trial][0] == '1':
            reinforce_counter += 1
    
    wb = openpyxl.Workbook() # Create a excel Workbook 
    ws = wb.active # Create a excel worksheet
    ws.title = "Duração da sequência"
    ws.append(["Duração da primeira, segunda, penúltima e última sequência"])
    ws.append(["Reforço", "Primeira", "Segunda", "Penúltima", "Ultima"])
    all_relevant_duration = [],[],[],[]
    for i in duration_by_reinforce:
        all_relevant_duration[0].append(duration_by_reinforce[i][0])
        all_relevant_duration[1].append(duration_by_reinforce[i][1])
        all_relevant_duration[2].append(duration_by_reinforce[i][-2])
        all_relevant_duration[3].append(duration_by_reinforce[i][-1])
       
        ws.append([f"{i}", duration_by_reinforce[i][0], duration_by_reinforce[i][1], duration_by_reinforce[i][-2], duration_by_reinforce[i][-1]])


    ws.append(["Média", statistics.mean(all_relevant_duration[0]), statistics.mean(all_relevant_duration[1]), statistics.mean(all_relevant_duration[2]), statistics.mean(all_relevant_duration[3])])
    ws.append([])
    ws.append([])


    correct_duration = [[]]
    reinforce_counter = 0

    for i in range(len(row_consequence_data)):
        if row_consequence_data[i][0] == '1':
            correct_duration[reinforce_counter].append(time_data_separated[i])
            reinforce_counter += 1
        if row_consequence_data[i][0] == '2':
            try:
                correct_duration[reinforce_counter].append(time_data_separated[i])
            except IndexError:
                correct_duration.append([time_data_separated[i]])

    ws.append(["Duração das sequências corretas"])
    ws.append(["Reforço", "Primeira", "Segunda", "Terceira", "Quarta"])
    temp_list = [[] for i in range(4)]
    for i in range(len(correct_duration)):
        temp = [f"{i + 1}"]
        for j in range(len(correct_duration[i])):
            try:
                temp_list[j].append(correct_duration[i][j])
                temp.append(correct_duration[i][j])
            except IndexError:
                pass
        ws.append(temp)
    
    ws.append(["Média", statistics.mean(temp_list[0]), statistics.mean(temp_list[1]), statistics.mean(temp_list[2]), statistics.mean(temp_list[3])])

    # Coping data to excel
    wb.save(f"./planilhas/{name}.xlsx")

def calc_trial_duration(time_data, consequence_data, individual, name):
    cb = QClipboard()
    row_time_data = clear_data(time_data)
    row_consequence_data = clear_data(consequence_data)

    time_data_separated = []
    for i in row_time_data:
        time_data_separated.append(int(i.split('.')[0]))
   
    reinforce_counter = 1
    duration_by_reinforce = {}

    for trial in range(len(row_consequence_data)):
        if reinforce_counter in duration_by_reinforce.keys():
            duration_by_reinforce[reinforce_counter].append(time_data_separated[trial])
        else:
            duration_by_reinforce[reinforce_counter] = [time_data_separated[trial]]
    
        if row_consequence_data[trial][0] == '1':
            reinforce_counter += 1
    
    wb = openpyxl.Workbook() # Create a excel Workbook 
    ws = wb.active # Create a excel worksheet
    ws.title = "Duração da tentativa"
    ws.append(["Duração da primeira, segunda, penúltima e última tentativa"])
    ws.append(["Reforço", "Primeira", "Segunda", "Penúltima", "Ultima"])
    all_relevant_duration = [],[],[],[]
    for i in duration_by_reinforce:
        temp = [f"{i}"]
        for j in (0, 1, -2, -1):
            try:
                temp.append(duration_by_reinforce[i][j])
                all_relevant_duration[j].append(duration_by_reinforce[i][j])
            except IndexError:
                pass
       
        ws.append(temp)


    ws.append(["Média", statistics.mean(all_relevant_duration[0]), statistics.mean(all_relevant_duration[1]), statistics.mean(all_relevant_duration[2]), statistics.mean(all_relevant_duration[3])])
    ws.append([])
    ws.append([])


    correct_duration = [[]]
    reinforce_counter = 0

    for i in range(len(row_consequence_data)):
        if row_consequence_data[i][0] == '1':
            correct_duration[reinforce_counter].append(time_data_separated[i])
            reinforce_counter += 1
        if row_consequence_data[i][0] == '2':
            try:
                correct_duration[reinforce_counter].append(time_data_separated[i])
            except IndexError:
                correct_duration.append([time_data_separated[i]])

    if individual:
        ws.append(["Duração das respostas corretas"])
        ws.append(["Reforço", "Primeira", "Segunda", "Terceira", "Quarta", "Quinta", "Sexta", "Sétima", "Oitava", "Nona", "Decima", "Decima Primeira", "Decima Segunda"])

        temp_list = [[] for i in range(12)]
        for i in range(len(correct_duration)):
            temp = [f"{i + 1}"]
            for j in range(len(correct_duration[i])):
                temp_list[j].append(correct_duration[i][j])
                temp.append(correct_duration[i][j])
            ws.append(temp)
        temp_mean = ["Média:"]
        for i in range(len(temp_list)):
            temp_mean.append(statistics.mean(temp_list[i]))
        ws.append(temp_mean)

    else:
        ws.append(["Duração das tentativas corretas"])
        ws.append(["Reforço", "Primeira", "Segunda", "Terceira", "Quarta"])
        temp_list = [[] for i in range(4)]
        for i in range(len(correct_duration)):
            temp = [f"{i + 1}",]
            for j in range(len(correct_duration[i])):
                temp_list[j].append(correct_duration[i][j])
                temp.append(correct_duration[i][j])
            ws.append(temp)
        
        ws.append(["Média", statistics.mean(temp_list[0]), statistics.mean(temp_list[1]), statistics.mean(temp_list[2]), statistics.mean(temp_list[3])])

    # Coping data to excel
    wb.save(f"./planilhas/{name}.xlsx")
