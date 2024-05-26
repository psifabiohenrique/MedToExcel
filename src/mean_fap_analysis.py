from PySide6.QtGui import QClipboard
import openpyxl
import statistics
from src.Recorrence import remover_data



def clear_data(string):

    row_data = string.split()
    if ',' in string:
        temp = []
        for i in row_data:
            temp.append(i.replace(',', '.'))
        row_data = temp
    if '/' in string:
        row_data = remover_data(row_data)
    
    temp = []
    for i in row_data:
        if '0.000' != i:
            temp.append(i)

    row_data = temp
    return row_data

def separate_data(time_data): 
        all_latency = []
        all_f2s = []
        all_s2t = []
        for i in time_data:
            if i[-3:] == '001':
                temp = i[:-3]
                temp = temp.replace('.', '')
                all_latency.append(int(temp))
            elif i[-3:] == '002':
                temp = i[:-3]
                temp = temp.replace('.', '')
                all_f2s.append(int(temp))
            elif i[-3:] == '003':
                temp = i[:-3]
                temp = temp.replace('.', '')
                all_s2t.append(int(temp))
        
        return (all_latency, all_f2s, all_s2t)

def calc_latency(time_data, consequence_data, individual, name):
    cb = QClipboard()
    row_time_data = clear_data(time_data)
    row_consequence_data = clear_data(consequence_data)

    if individual:
        time_data_separated = []
        for i in row_time_data:
            time_data_separated.append(int(i.split('.')[0]))
    else:
        time_data_separated = separate_data(row_time_data)
    
    reinforce_counter = 1
    latency_by_reinforce = {}
    f2s_by_reinforce = {}
    s2t_by_reinforce = {}

    for trial in range(len(row_consequence_data)):
        if reinforce_counter in latency_by_reinforce.keys():
            if individual:
                latency_by_reinforce[reinforce_counter].append(time_data_separated[trial])
            else:
                latency_by_reinforce[reinforce_counter].append(time_data_separated[0][trial])
                f2s_by_reinforce[reinforce_counter].append(time_data_separated[1][trial])
                s2t_by_reinforce[reinforce_counter].append(time_data_separated[2][trial])
        else:
            if individual:
                latency_by_reinforce[reinforce_counter] = [time_data_separated[trial]]
            else:
                latency_by_reinforce[reinforce_counter] = [time_data_separated[0][trial]]
                f2s_by_reinforce[reinforce_counter] = [time_data_separated[1][trial]]
                s2t_by_reinforce[reinforce_counter] = [time_data_separated[2][trial]]

        if row_consequence_data[trial][0] == '1':
            reinforce_counter += 1
    
    wb = openpyxl.Workbook() # Create a excel Workbook 
    ws = wb.active # Create a excel worksheet
    ws.title = "Latencia"
    ws.append(["Latência da primeira, segunda, penúltima e última sequência"])
    ws.append(["Reforço", "Primeira", "Segunda", "Penúltima", "Ultima"])
    all_relevant_latency = [],[],[],[]
    for i in latency_by_reinforce:
        all_relevant_latency[0].append(latency_by_reinforce[i][0])
        all_relevant_latency[1].append(latency_by_reinforce[i][1])
        all_relevant_latency[2].append(latency_by_reinforce[i][-2])
        all_relevant_latency[3].append(latency_by_reinforce[i][-1])
       
        ws.append([f"{i}", latency_by_reinforce[i][0], latency_by_reinforce[i][1], latency_by_reinforce[i][-2], latency_by_reinforce[i][-1]])


    ws.append(["Média", statistics.mean(all_relevant_latency[0]), statistics.mean(all_relevant_latency[1]), statistics.mean(all_relevant_latency[2]), statistics.mean(all_relevant_latency[3])])
    ws.append([])
    ws.append([])

    if not individual:
        ws2 = wb.create_sheet("Primeira para Segunda") # Create a excel worksheet
        ws2.append(["Tempo entre a primeira e a segunda resposta da primeira, segunda, penúltima e última sequência"])
        ws2.append(["Reforço", "Primeira", "Segunda", "Penúltima", "Ultima"])
        all_relevant_f2s = [],[],[],[]
        for i in f2s_by_reinforce:
            all_relevant_f2s[0].append(f2s_by_reinforce[i][0])
            all_relevant_f2s[1].append(f2s_by_reinforce[i][1])
            all_relevant_f2s[2].append(f2s_by_reinforce[i][-2])
            all_relevant_f2s[3].append(f2s_by_reinforce[i][-1])
        
            ws2.append([f"{i}", f2s_by_reinforce[i][0], f2s_by_reinforce[i][1], f2s_by_reinforce[i][-2], f2s_by_reinforce[i][-1]])

        ws2.append(["Média", statistics.mean(all_relevant_f2s[0]), statistics.mean(all_relevant_f2s[1]), statistics.mean(all_relevant_f2s[2]), statistics.mean(all_relevant_f2s[3])])
        ws2.append([])
        ws2.append([])


        ws3 = wb.create_sheet("Segunda para Terceira") # Create a excel worksheet
        ws3.append(["Tempo entre a primeira e a segunda resposta da primeira, segunda, penúltima e última sequência"])
        ws3.append(["Reforço", "Primeira", "Segunda", "Penúltima", "Ultima"])
        all_relevant_s2t = [],[],[],[]
        for i in s2t_by_reinforce:
            all_relevant_s2t[0].append(s2t_by_reinforce[i][0])
            all_relevant_s2t[1].append(s2t_by_reinforce[i][1])
            all_relevant_s2t[2].append(s2t_by_reinforce[i][-2])
            all_relevant_s2t[3].append(s2t_by_reinforce[i][-1])
        
            ws3.append([f"{i}", s2t_by_reinforce[i][0], s2t_by_reinforce[i][1], s2t_by_reinforce[i][-2], s2t_by_reinforce[i][-1]])

        ws3.append(["Média", statistics.mean(all_relevant_s2t[0]), statistics.mean(all_relevant_s2t[1]), statistics.mean(all_relevant_s2t[2]), statistics.mean(all_relevant_s2t[3])])
        ws3.append([])
        ws3.append([])

    # Coping data to excel
    wb.save(f"./planilhas/{name}.xlsx")


def calc_sequence_duration(time_data, consequence_data, name):
    cb = QClipboard()
    row_time_data = clear_data(time_data)
    row_consequence_data = clear_data(consequence_data)

    time_data_separated = []
    for i in row_time_data:
        time_data_separated.append(int(i.split('.')[0]))
   
    reinforce_counter = 1
    duration_by_reinforce = {}

    for trial in range(len(row_consequence_data)):
        if reinforce_counter in duration_by_reinforce.keys():
            duration_by_reinforce[reinforce_counter].append(time_data_separated[trial])
        else:
            duration_by_reinforce[reinforce_counter] = [time_data_separated[trial]]
    
        if row_consequence_data[trial][0] == '1':
            reinforce_counter += 1
    
    wb = openpyxl.Workbook() # Create a excel Workbook 
    ws = wb.active # Create a excel worksheet
    ws.title = "Duração da sequência"
    ws.append(["Duração da primeira, segunda, penúltima e última sequência"])
    ws.append(["Reforço", "Primeira", "Segunda", "Penúltima", "Ultima"])
    all_relevant_duration = [],[],[],[]
    for i in duration_by_reinforce:
        all_relevant_duration[0].append(duration_by_reinforce[i][0])
        all_relevant_duration[1].append(duration_by_reinforce[i][1])
        all_relevant_duration[2].append(duration_by_reinforce[i][-2])
        all_relevant_duration[3].append(duration_by_reinforce[i][-1])
       
        ws.append([f"{i}", duration_by_reinforce[i][0], duration_by_reinforce[i][1], duration_by_reinforce[i][-2], duration_by_reinforce[i][-1]])


    ws.append(["Média", statistics.mean(all_relevant_duration[0]), statistics.mean(all_relevant_duration[1]), statistics.mean(all_relevant_duration[2]), statistics.mean(all_relevant_duration[3])])
    ws.append([])
    ws.append([])

    # Coping data to excel
    wb.save(f"./planilhas/{name}.xlsx")

def calc_trial_duration(time_data, consequence_data, name):
    cb = QClipboard()
    row_time_data = clear_data(time_data)
    row_consequence_data = clear_data(consequence_data)

    time_data_separated = []
    for i in row_time_data:
        time_data_separated.append(int(i.split('.')[0]))
   
    reinforce_counter = 1
    duration_by_reinforce = {}

    for trial in range(len(row_consequence_data)):
        if reinforce_counter in duration_by_reinforce.keys():
            duration_by_reinforce[reinforce_counter].append(time_data_separated[trial])
        else:
            duration_by_reinforce[reinforce_counter] = [time_data_separated[trial]]
    
        if row_consequence_data[trial][0] == '1':
            reinforce_counter += 1
    
    wb = openpyxl.Workbook() # Create a excel Workbook 
    ws = wb.active # Create a excel worksheet
    ws.title = "Duração da tentativa"
    ws.append(["Duração da primeira, segunda, penúltima e última tentativa"])
    ws.append(["Reforço", "Primeira", "Segunda", "Penúltima", "Ultima"])
    all_relevant_duration = [],[],[],[]
    for i in duration_by_reinforce:
        all_relevant_duration[0].append(duration_by_reinforce[i][0])
        all_relevant_duration[1].append(duration_by_reinforce[i][1])
        all_relevant_duration[2].append(duration_by_reinforce[i][-2])
        all_relevant_duration[3].append(duration_by_reinforce[i][-1])
       
        ws.append([f"{i}", duration_by_reinforce[i][0], duration_by_reinforce[i][1], duration_by_reinforce[i][-2], duration_by_reinforce[i][-1]])


    ws.append(["Média", statistics.mean(all_relevant_duration[0]), statistics.mean(all_relevant_duration[1]), statistics.mean(all_relevant_duration[2]), statistics.mean(all_relevant_duration[3])])
    ws.append([])
    ws.append([])

    # Coping data to excel
    wb.save(f"./planilhas/{name}.xlsx")
